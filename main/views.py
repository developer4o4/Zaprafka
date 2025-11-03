from decimal import Decimal
from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login,logout
from .models import *
# Create your views here.
@login_required
def home(request):
    """Bosh sahifa uchun bugungi statistikani qaytarish"""
    today = timezone.now().date()
    tashkilotlar = Tashkilot.objects.all()
    # Bugungi ma'lumotlar
    today_data = Compilated.objects.filter(created_ad__date=today,who_user=request.user)
    
    # Statistik ma'lumotlar
    total_fuel = today_data.aggregate(total=Sum('hajm'))['total'] or 0
    total_count = today_data.count()
    active_cars = today_data.values('avto').distinct().count()
    avg_fuel = total_fuel / total_count if total_count > 0 else 0
    
    # So'nggi 5 ta faoliyat
    recent_activities = today_data.select_related('tashkilot', 'avto').order_by('-created_ad')[:5]
    
    activities_list = []
    for activity in recent_activities:
        activities_list.append({
            'avto_title': activity.avto.title,
            'avto_number': activity.avto.avto_number,
            'tashkilot_title': activity.tashkilot.title,
            'hajm': activity.hajm,
            'time': activity.created_ad.strftime('%H:%M')
        })
    
    response_data = {
        'total_fuel': round(total_fuel, 1),
        'total_count': total_count,
        "tashkilotlar":tashkilotlar,

        'active_cars': active_cars,
        'avg_fuel': round(avg_fuel, 1),
        'recent_activities': activities_list
    }
    
    return render(request, "home.html",response_data)


from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login

def login_view(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is None:
            return redirect('login')

        login(request, user)
        if user.is_superuser:
            return redirect('admin_panel')
        return redirect("home")

    else:
        if request.user.is_authenticated:
            return redirect('home')  # kerakli sahifani yozing
        else:
            return render(request, "login.html", {"error": "Parol yoki username hato!"})

@login_required
# def admin_panel(request):
# 	tashkilotlar = Tashkilot.objects.all()
# 	avto_list = Avto.objects.all()
# 	yoqilgilar = Yoqilgi_turi.objects.all()
# 	return render(request, "admin.html", {
# 	"tashkilotlar": tashkilotlar,
# 	"avto_list": avto_list,
# 	"yoqilgilar": yoqilgilar,
# 	})
@login_required
def logout_view(request):
	logout(request)
	return redirect("login")


@login_required
def add_user(request):
	if not request.user.is_superuser:
		return render(request, "add_user.html", {"error": "Sizda foydalanuvchi yaratish huquqi yo‘q."})

	context = {}

	if request.method == 'POST':
		username = request.POST.get('username')
		phone = request.POST.get('tel')
		password = request.POST.get('password')

		if not username or not password:
			context["error"] = "Foydalanuvchi nomi va parol to‘ldirilishi shart."
			return render(request, "add_user.html", context)

		if User.objects.filter(username=username).exists():
			context["error"] = "Bu foydalanuvchi nomi allaqachon mavjud."
			return render(request, "add_user.html", context)

		user = User.objects.create_user(username=username, phone=phone, password=password)
		user.is_staff = False
		user.is_superuser = False
		user.save()

		context["success"] = f"✅ {username} foydalanuvchisi muvaffaqiyatli yaratildi."
		return render(request, "add_user.html", context)

	return render(request, "add_user.html")
@login_required
def user_delete(request, pk):
    """User ni o'chirish"""
    user_to_delete = get_object_or_404(User, pk=pk)
    
    # O'zini o'chirishni oldini olish
    if request.user == user_to_delete:
        messages.error(request, 'Siz o\'zingizni o\'chira olmaysiz!')
        return redirect('admin_panel')
    
    if request.method == 'POST':
        try:
            username = user_to_delete.username
            user_to_delete.delete()
            messages.success(request, f'Foydalanuvchi "{username}" muvaffaqiyatli o\'chirildi!')
            return redirect('admin_panel')
        except Exception as e:
            messages.error(request, f'Foydalanuvchini o\'chirishda xatolik: {str(e)}')
            return redirect('admin_panel')
    
    context = {
        'user_to_delete': user_to_delete
    }
    return render(request, 'user_delete.html', context)
@login_required
def add_tashkilot(request):
    if request.method == "POST":
        title = request.POST.get("title")
        group_id = request.POST.get("group_id")
        if title:
            Tashkilot.objects.create(title=title, group_id=group_id)
            return redirect("admin_panel")
    return render(request, "add_tashkilot.html")

@login_required
def add_avto(request):
    tashkilotlar = Tashkilot.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title')
        avto_number = request.POST.get('avto_number')
        tashkilot_id = request.POST.get('tashkilot_id')

        tashkilot = Tashkilot.objects.filter(id=tashkilot_id).first()
        if tashkilot:
            Avto.objects.create(
                title=title,
                avto_number=avto_number,
                tashkilot=tashkilot
            )
            # shu joyda render qaytariladi, redirect emas
            return redirect('admin_panel')
            # return render(request, 'admin.html', {
            #     'tashkilotlar': tashkilotlar,
            #     'success': "Avto muvaffaqiyatli qo‘shildi!"
            # })
        else:
            return render(request, 'admin.html', {
                'tashkilotlar': tashkilotlar,
                'error': "Tashkilot topilmadi!"
            })

    return render(request, 'admin.html', {
        'tashkilotlar': tashkilotlar
    })

@login_required
def add_yoqilgi(request):
	if request.method == "POST":
		title = request.POST.get("title")
		price = request.POST.get("price")
		if title:
			Yoqilgi_turi.objects.create(title=title,price=price)
			return redirect("admin_panel")
	return render(request, "add_yoqilgi.html")
from django.db.models import Sum, Count, Avg, Max, Min
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
import json
from .models import Compilated, Tashkilot, Yoqilgi_turi, Avto
@login_required
def statistics_view(request):
    """Statistika sahifasini ko'rsatish"""
    try:
        tashkilotlar = Tashkilot.objects.all()
        yoqilgi_turlari = Yoqilgi_turi.objects.all()
        
        context = {
            'tashkilotlar': tashkilotlar,
            'yoqilgi_turlari': yoqilgi_turlari,
        }
        return render(request, 'stats.html', context)
    except Exception as e:
        # Xatolik yuz bersa, bo'sh kontekst bilan sahifani ko'rsatish
        return render(request, 'stats.html', {})
@login_required
def get_statistics_data(request):
    """Statistika ma'lumotlarini JSON formatida qaytarish"""
    try:
        # Filtrlarni olish
        period = request.GET.get('period', '30')
        tashkilot_id = request.GET.get('tashkilot', 'all')
        yoqilgi_id = request.GET.get('yoqilgi', 'all')
        detail_period = request.GET.get('detail_period', 'daily')
        
        # Periodni tekshirish va konvertatsiya qilish
        try:
            period_days = int(period)
            if period_days <= 0:
                period_days = 30
        except (ValueError, TypeError):
            period_days = 30
        
        # Sana oralig'ini belgilash
        end_date = timezone.now()
        start_date = end_date - timedelta(days=period_days)
        
        # Compilated ma'lumotlarini filtrlash
        compilated_data = Compilated.objects.filter(
            created_ad__gte=start_date,
            created_ad__lte=end_date
        )
        
        # Tashkilot bo'yicha filtr
        if tashkilot_id != 'all':
            try:
                compilated_data = compilated_data.filter(tashkilot_id=int(tashkilot_id))
            except (ValueError, TypeError):
                pass  # Noto'g'ri ID berilsa, filtrni qo'llamaymiz
        
        # Yoqilgi turi bo'yicha filtr
        if yoqilgi_id != 'all':
            try:
                compilated_data = compilated_data.filter(yoqilgi_turi_id=int(yoqilgi_id))
            except (ValueError, TypeError):
                pass  # Noto'g'ri ID berilsa, filtrni qo'llamaymiz
        
        # Ma'lumotlar mavjudligini tekshirish
        if not compilated_data.exists():
            return JsonResponse({
                'summary': {
                    'total_fuel': 0,
                    'avg_daily': 0,
                    'max_fuel': 0,
                    'min_fuel': 0,
                    'active_cars': 0,
                    'total_records': 0
                },
                'tashkilot_stats': [],
                'avto_stats': [],
                'daily_stats': [],
                'monthly_stats': [],
                'weekly_stats': [],
                'detailed_stats': [],
                'recent_records': []
            })
        
        # Umumiy statistik ma'lumotlar
        aggregates = compilated_data.aggregate(
            total=Sum('hajm'),
            avg=Avg('hajm'),
            max=Max('hajm'),
            min=Min('hajm'),
            count=Count('id')
        )
        
        total_fuel = aggregates['total'] or 0
        avg_daily = total_fuel / period_days if period_days > 0 else 0
        max_fuel = aggregates['max'] or 0
        min_fuel = aggregates['min'] or 0
        active_cars = compilated_data.values('avto').distinct().count()
        total_records = aggregates['count'] or 0
        
        # Tashkilotlar bo'yicha ma'lumotlar
        tashkilot_stats = list(compilated_data.values(
            'tashkilot__id',
            'tashkilot__title'
        ).annotate(
            total_fuel=Sum('hajm'),
            record_count=Count('id'),
            avg_fuel=Avg('hajm')
        ).order_by('-total_fuel'))
        
        # Avtomobillar bo'yicha ma'lumotlar
        avto_stats = list(compilated_data.values(
            'avto__id',
            'avto__title',
            'avto__avto_number'
        ).annotate(
            total_fuel=Sum('hajm'),
            record_count=Count('id'),
            avg_fuel=Avg('hajm')
        ).order_by('-total_fuel')[:10])  # Faqat top 10
        
        # Kunlik ma'lumotlar
        daily_stats = list(compilated_data.annotate(
            day=TruncDay('created_ad')
        ).values('day').annotate(
            daily_total=Sum('hajm'),
            record_count=Count('id')
        ).order_by('day'))
        
        # Oylik ma'lumotlar
        monthly_stats = list(compilated_data.annotate(
            month=TruncMonth('created_ad')
        ).values('month').annotate(
            monthly_total=Sum('hajm'),
            record_count=Count('id')
        ).order_by('month'))
        
        # Haftalik ma'lumotlar
        weekly_stats = list(compilated_data.annotate(
            week=TruncWeek('created_ad')
        ).values('week').annotate(
            weekly_total=Sum('hajm'),
            record_count=Count('id')
        ).order_by('week'))
        
        # Batafsil statistika
        if detail_period == 'daily':
            detailed_stats = compilated_data.annotate(
                period=TruncDay('created_ad')
            ).values('period').annotate(
                total_fuel=Sum('hajm'),
                record_count=Count('id'),
                tashkilot_count=Count('tashkilot', distinct=True),
                avto_count=Count('avto', distinct=True),
                avg_fuel=Avg('hajm')
            ).order_by('-period')
        elif detail_period == 'weekly':
            detailed_stats = compilated_data.annotate(
                period=TruncWeek('created_ad')
            ).values('period').annotate(
                total_fuel=Sum('hajm'),
                record_count=Count('id'),
                tashkilot_count=Count('tashkilot', distinct=True),
                avto_count=Count('avto', distinct=True),
                avg_fuel=Avg('hajm')
            ).order_by('-period')
        else:  # monthly
            detailed_stats = compilated_data.annotate(
                period=TruncMonth('created_ad')
            ).values('period').annotate(
                total_fuel=Sum('hajm'),
                record_count=Count('id'),
                tashkilot_count=Count('tashkilot', distinct=True),
                avto_count=Count('avto', distinct=True),
                avg_fuel=Avg('hajm')
            ).order_by('-period')
        
        # So'nggi 10 ta yoqilg'i quyish ma'lumotlari
        recent_records = compilated_data.select_related(
            'tashkilot', 'avto', 'who_user'
        ).order_by('-created_ad')[:10]
        
        recent_records_list = []
        for record in recent_records:
            photo_url = None
            if record.photo and hasattr(record.photo, 'url') and record.photo.name != 'none.jpg':
                try:
                    photo_url = request.build_absolute_uri(record.photo.url)
                except:
                    photo_url = None
            
            recent_records_list.append({
                'id': record.id,
                'tashkilot_title': record.tashkilot.title if record.tashkilot else 'Noma\'lum',
                'avto_title': record.avto.title if record.avto else 'Noma\'lum',
                'avto_number': record.avto.avto_number if record.avto else '',
                'user_name': record.who_user.username if record.who_user else 'Noma\'lum',
                'hajm': float(record.hajm) if record.hajm else 0,
                'yoqilgi_turi': record.yoqilgi_turi if record.yoqilgi_turi else 'Noma\'lum',
                'all_price': float(record.all_price) if record.all_price else 0,
                'created_ad': record.created_ad.strftime('%Y-%m-%d %H:%M') if record.created_ad else 'Noma\'lum',
                'photo_url': photo_url
            })
        
        # Ma'lumotlarni formatlash va yuborish
        response_data = {
            'success': True,
            'summary': {
                'total_fuel': round(float(total_fuel), 1),
                'avg_daily': round(float(avg_daily), 1),
                'max_fuel': round(float(max_fuel), 1),
                'min_fuel': round(float(min_fuel), 1),
                'active_cars': active_cars,
                'total_records': total_records,
                'period_days': period_days
            },
            'tashkilot_stats': tashkilot_stats,
            'avto_stats': avto_stats,
            'daily_stats': daily_stats,
            'monthly_stats': monthly_stats,
            'weekly_stats': weekly_stats,
            'detailed_stats': list(detailed_stats),
            'recent_records': recent_records_list
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        # Xatolik yuz bersa, xabar bilan javob qaytarish
        return JsonResponse({
            'success': False,
            'error': str(e),
            'summary': {
                'total_fuel': 0,
                'avg_daily': 0,
                'max_fuel': 0,
                'min_fuel': 0,
                'active_cars': 0,
                'total_records': 0
            },
            'tashkilot_stats': [],
            'avto_stats': [],
            'daily_stats': [],
            'monthly_stats': [],
            'weekly_stats': [],
            'detailed_stats': [],
            'recent_records': []
        }, status=500)
@login_required
def get_filter_options(request):
    """Filtrlash uchun variantlarni qaytarish"""
    try:
        tashkilotlar = list(Tashkilot.objects.all().values('id', 'title').order_by('title'))
        yoqilgi_turlari = list(Yoqilgi_turi.objects.all().values('id', 'title').order_by('title'))
        avtomobillar = list(Avto.objects.select_related('tashkilot').values(
            'id', 'title', 'avto_number', 'tashkilot_id'
        ).order_by('title'))
        
        response_data = {
            'success': True,
            'tashkilotlar': tashkilotlar,
            'yoqilgi_turlari': yoqilgi_turlari,
            'avtomobillar': avtomobillar
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'tashkilotlar': [],
            'yoqilgi_turlari': [],
            'avtomobillar': []
        }, status=500)
@login_required
def export_statistics_excel(request):
    """Statistika ma'lumotlarini Excel fayl sifatida eksport qilish"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        
        # Filtrlarni olish (statistika ma'lumotlari bilan bir xil)
        period = request.GET.get('period', '30')
        tashkilot_id = request.GET.get('tashkilot', 'all')
        yoqilgi_id = request.GET.get('yoqilgi', 'all')
        
        # Sana oralig'ini belgilash
        try:
            period_days = int(period)
            if period_days <= 0:
                period_days = 30
        except (ValueError, TypeError):
            period_days = 30
            
        end_date = timezone.now()
        start_date = end_date - timedelta(days=period_days)
        
        # Ma'lumotlarni olish
        compilated_data = Compilated.objects.filter(
            created_ad__gte=start_date,
            created_ad__lte=end_date
        ).select_related('tashkilot', 'avto', 'who_user')
        
        if tashkilot_id != 'all':
            try:
                compilated_data = compilated_data.filter(tashkilot_id=int(tashkilot_id))
            except (ValueError, TypeError):
                pass
        
        if yoqilgi_id != 'all':
            try:
                compilated_data = compilated_data.filter(yoqilgi_turi_id=int(yoqilgi_id))
            except (ValueError, TypeError):
                pass
        
        # Excel fayl yaratish
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"yoqilgi_statistikasi_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ma'lumotlarni DataFrame ga o'tkazish
        data_list = []
        for record in compilated_data:
            data_list.append({
                'Sana': record.created_ad.strftime('%Y-%m-%d %H:%M'),
                'Tashkilot': record.tashkilot.title if record.tashkilot else '',
                'Avtomobil': record.avto.title if record.avto else '',
                'Avtomobil raqami': record.avto.avto_number if record.avto else '',
                'Yoqilg\'i turi': record.yoqilgi_turi,
                'Miqdor (L)': float(record.hajm) if record.hajm else 0,
                'Jami narx': float(record.all_price) if record.all_price else 0,
                'Foydalanuvchi': record.who_user.username if record.who_user else '',
            })
        
        df = pd.DataFrame(data_list)
        
        # Excel faylini yaratish
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            # Asosiy ma'lumotlar
            if not df.empty:
                df.to_excel(writer, sheet_name='Yoqilg\'i Quyishlar', index=False)
            
            # Statistika jadvali
            summary_data = {
                'Ko\'rsatkich': [
                    'Jami Yoqilg\'i (L)',
                    'O\'rtacha Kunlik (L)', 
                    'Maksimal (L)',
                    'Avtomobillar Soni',
                    'Yozuvlar Soni',
                    'Davr (kun)'
                ],
                'Qiymat': [
                    round(df['Miqdor (L)'].sum(), 1) if not df.empty else 0,
                    round(df['Miqdor (L)'].sum() / period_days, 1) if not df.empty and period_days > 0 else 0,
                    round(df['Miqdor (L)'].max(), 1) if not df.empty else 0,
                    df['Avtomobil'].nunique() if not df.empty else 0,
                    len(df),
                    period_days
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Statistika', index=False)
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Excel fayl yaratishda xatolik: {str(e)}'
        }, status=500)
@login_required
def yoqilgi_quyish(request):
    """Yoqilg'i qo'shish sahifasi"""
    tashkilotlar = Tashkilot.objects.all()
    yoqilgi_turlari = Yoqilgi_turi.objects.all()
    
    context = {
        'tashkilotlar': tashkilotlar,
        'yoqilgi_turlari': yoqilgi_turlari,
    }
    return render(request, 'yoqilgi_quyish.html', context)
@login_required
def bugungi_yoqilgilar(request):
    """Bugungi yoqilg'ilar sahifasi"""
    tashkilotlar = Tashkilot.objects.all()
    yoqilgi_turlari = Yoqilgi_turi.objects.all()
    
    context = {
        'tashkilotlar': tashkilotlar,
        'yoqilgi_turlari': yoqilgi_turlari,
    }
    return render(request, 'bugungi_yoqilgilar.html', context)

import requests
from django.db.models.functions import ExtractHour
from django.http import HttpResponse, JsonResponse
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.base import ContentFile
import base64
import datetime
@csrf_exempt
def send_telegram(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            group_id = data.get('group_id')
            message = data.get('message')
            process_photo = data.get('process_photo')
            confirmation_photo = data.get('confirmation_photo')
            
            # Telegram bot tokeni
            bot_token = '8384548755:AAE_O3g_2Q971QHNU8eqk3NCo7bxTAZrf9o'
            
            # Chat ID ni formatlash
            chat_id = f"-{group_id}"
            
            # 1. Avval matnli xabarni yuborish
            url = f'https://api.telegram.org/bot{bot_token}/sendMessage'
            payload = {
                'chat_id': chat_id,
                'text': message,
                'parse_mode': 'HTML'
            }
            
            # response = requests.post(url, json=payload)
            # print(f"Matn xabari status: {response.status_code}")
            
            # 2. Ikkala rasmni albom shaklida yuborish
            def create_media_group(photos_data, caption):
                media = []
                
                for i, photo_data in enumerate(photos_data):
                    if photo_data and photo_data != '':
                        try:
                            # Base64 rasmni faylga aylantirish
                            photo_data_clean = photo_data.split(',')[1] if ',' in photo_data else photo_data
                            photo_bytes = base64.b64decode(photo_data_clean)
                            
                            # Rasmni serverga vaqtincha yuklash
                            photo_file = {
                                'photo': ('photo.jpg', photo_bytes, 'image/jpeg')
                            }
                            
                            # Har bir rasm uchun media ob'ekt yaratish
                            media.append({
                                'type': 'photo',
                                'media': f'attach://photo_{i}',
                                'caption': caption if i == 0 else ''  # Faqat birinchi rasmga caption
                            })
                            
                        except Exception as e:
                            print(f"Rasm tayyorlash xatosi: {e}")
                            continue
                
                return media
            
            # Rasmlar ro'yxatini tayyorlash
            photos_data = []
            captions = []
            
            if process_photo and process_photo != '':
                photos_data.append(process_photo)
                captions.append("🛢️ Yoqilg'i quyish jarayoni")
            
            if confirmation_photo and confirmation_photo != '':
                photos_data.append(confirmation_photo)
                captions.append("✅ Yoqilg'i quyilganidan keyingi holat")
            
            # Agar ikkala rasm ham mavjud bo'lsa, albom shaklida yuborish
            if len(photos_data) >= 1:
                if len(photos_data) == 2:
                    print("ikkita")
                    # Ikkala rasmni albom shaklida yuborish
                    url = f'https://api.telegram.org/bot{bot_token}/sendMediaGroup'
                    
                    media = []
                    files = {}
                    
                    for i, photo_data in enumerate(photos_data):
                        if photo_data and photo_data != '':
                            try:
                                # Base64 rasmni faylga aylantirish
                                photo_data_clean = photo_data.split(',')[1] if ',' in photo_data else photo_data
                                photo_bytes = base64.b64decode(photo_data_clean)
                                
                                # Fayl nomi
                                file_key = f'photo_{i}'
                                files[file_key] = (f'photo_{i}.jpg', photo_bytes, 'image/jpeg')
                                
                                # Media ob'ekt
                                media_item = {
                                    'type': 'photo',
                                    'media': f'attach://{file_key}'
                                }
                                
                                # Faqat birinchi rasmga caption qo'shish
                                if i == 0:
                                    media_item['caption'] = message
                                
                                media.append(media_item)
                                
                            except Exception as e:
                                print(f"Rasm {i} tayyorlash xatosi: {e}")
                                continue
                    
                    if media:
                        payload = {
                            'chat_id': chat_id,
                            'media': json.dumps(media)
                        }
                        
                        response = requests.post(url, data=payload, files=files)
                        print(f"Albom status: {response.status_code}")
                        
                else:
                    # Faqat bitta rasm bo'lsa, oddiy yuborish
                    url = f'https://api.telegram.org/bot{bot_token}/sendPhoto'
                    
                    photo_data = photos_data[0]
                    photo_data_clean = photo_data.split(',')[1] if ',' in photo_data else photo_data
                    photo_bytes = base64.b64decode(photo_data_clean)
                    
                    files = {
                        'photo': ('photo.jpg', photo_bytes, 'image/jpeg')
                    }
                    data = {
                        'chat_id': chat_id,
                        'caption': message,
                        'parse_mode': 'HTML'
                    }
                    
                    response = requests.post(url, files=files, data=data)
                    print(f"Bitta rasm status: {response.status_code}")
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            print(f"Telegram xatosi: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Faqat POST so\'rovi qabul qilinadi'})
@login_required
def add_fuel(request):
    if request.method == 'POST':
        try:
            # Ma'lumotlarni olish
            tashkilot_id = request.POST.get('tashkilot')
            avto_id = request.POST.get('avtomobile')
            yoqilgi_id = request.POST.get('yoqilgi')
            miqdor = float(request.POST.get('miqdor'))
            all_price = Decimal(request.POST.get('all_price', 0))  # Decimal ga o'zgartirish
            captured_image = request.POST.get('captured_image')
            confirmation_photo = request.POST.get('confirmation_photo')
            yoqilgi_id = request.POST.get("yoqilgi")  # bu tanlangan ID
            if yoqilgi_id:
                yoqilgi_obj = Yoqilgi_turi.objects.get(id=yoqilgi_id)
                print(yoqilgi_obj)
            # Rasmni saqlash funksiyasi
            def save_base64_image(base64_string, filename):
                if base64_string and base64_string != '':
                    try:
                        format, imgstr = base64_string.split(';base64,')
                        ext = format.split('/')[-1]
                        data = ContentFile(base64.b64decode(imgstr), name=f'{filename}.{ext}')
                        return data
                    except Exception as e:
                        print(f"Rasm saqlash xatosi: {e}")
                        return None
                return None
            
            # Ma'lumotlarni saqlash
            compilated = Compilated(
                tashkilot_id=tashkilot_id,
                avto_id=avto_id,
                yoqilgi_turi=yoqilgi_obj,
                who_user=request.user,
                hajm=miqdor,
                all_price=all_price,
            )
            
            # Rasmlarni saqlash
            process_photo_file = save_base64_image(
                captured_image, 
                f'process_{request.user.id}_{int(timezone.now().timestamp())}'
            )
            if process_photo_file:
                compilated.photo = process_photo_file
            
            compilated.save()
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            print(f"Xatolik: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET so'rovi uchun
    tashkilotlar = Tashkilot.objects.all()
    yoqilgi_turlari = Yoqilgi_turi.objects.all()
    
    context = {
        'tashkilotlar': tashkilotlar,
        'yoqilgi_turlari': yoqilgi_turlari,
    }
    return render(request, 'add_fuel.html', context)
from datetime import date
@login_required
def get_today_fuel_api(request):
    """Bugungi yoqilg'ilarni olish API"""
    today = date.today()
    
    # Filtrlarni olish
    tashkilot_id = request.GET.get('tashkilot_id')
    avto_id = request.GET.get('avto_id')
    yoqilgi_id = request.GET.get('yoqilgi_id')
    
    # Bugungi ma'lumotlarni olish
    compilated_data = Compilated.objects.filter(
        created_ad__date=today
    ).select_related('tashkilot', 'avto', 'who_user')
    
    # Filtrlash
    if tashkilot_id and tashkilot_id != 'all':
        compilated_data = compilated_data.filter(tashkilot_id=tashkilot_id)
    
    if avto_id and avto_id != 'all':
        compilated_data = compilated_data.filter(avto_id=avto_id)
    
    # Yoqilg'i turi bo'yicha filtr (agar kerak bo'lsa)
    # if yoqilgi_id and yoqilgi_id != 'all':
    #     compilated_data = compilated_data.filter(yoqilgi_turi_id=yoqilgi_id)
    
    # Ma'lumotlarni tayyorlash
    records = []
    for record in compilated_data:
        records.append({
            'tashkilot_title': record.tashkilot.title,
            'avto_title': record.avto.title,
            'avto_number': record.avto.avto_number,
            'yoqilgi_turi': record.yoqilgi_turi,  # Bu yerda yoqilgi turini qo'shishingiz kerak
            'hajm': record.hajm,
            'created_ad': record.created_ad.isoformat(),
            'user_name': record.who_user.username
        })
    
    # Statistik ma'lumotlar
    total_fuel = sum(record.hajm for record in compilated_data)
    total_records = compilated_data.count()
    total_tashkilot = compilated_data.values('tashkilot').distinct().count()
    total_avto = compilated_data.values('avto').distinct().count()
    
    response_data = {
        'summary': {
            'total_fuel': round(total_fuel, 1),
            'total_records': total_records,
            'total_tashkilot': total_tashkilot,
            'total_avto': total_avto
        },
        'records': records
    }
    
    return JsonResponse(response_data)
@login_required
def get_avtomobillar_api(request):
    """Tashkilot bo'yicha avtomobillarni olish API"""
    tashkilot_id = request.GET.get('tashkilot_id')
    
    if tashkilot_id:
        avtomobillar = Avto.objects.filter(tashkilot_id=tashkilot_id)
    else:
        avtomobillar = Avto.objects.all()
    
    avto_list = []
    for avto in avtomobillar:
        avto_list.append({
            'id': avto.id,
            'title': avto.title,
            'avto_number': avto.avto_number
        })
    
    return JsonResponse(avto_list, safe=False)


from io import BytesIO
import pandas as pd
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Avg, Max, Min
from django.db.models.functions import TruncDay
from django.utils import timezone
from datetime import timedelta, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from decimal import Decimal
@login_required
def export_statistics_excel(request):
    """Statistika ma'lumotlarini Excel faylga eksport qilish"""
    try:
        # Filtrlarni olish
        period = request.GET.get('period', '30')
        tashkilot_id = request.GET.get('tashkilot', 'all')
        
        try:
            period_days = int(period)
        except ValueError:
            period_days = 30
        
        start_date = timezone.now() - timedelta(days=period_days)
        
        # Ma'lumotlarni olish
        compilated_data = Compilated.objects.filter(created_ad__gte=start_date)
        
        if tashkilot_id != 'all':
            compilated_data = compilated_data.filter(tashkilot_id=tashkilot_id)
        
        # Workbook yaratish
        wb = Workbook()
        
        # Agar hech qanday ma'lumot bo'lmasa
        if compilated_data.count() == 0:
            ws = wb.active
            ws.title = "Ma'lumot yo'q"
            ws['A1'] = "Tanlangan davrda ma'lumot topilmadi"
            ws['A2'] = f"Davr: So'nggi {period_days} kun"
            
            # Formatlash
            ws.column_dimensions['A'].width = 40
            ws['A1'].font = Font(bold=True, color='FF0000')
        else:
            # Formatlar
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            
            # 1. BARCHA MA'LUMOTLAR varaqasi
            ws_main = wb.active
            ws_main.title = "Barcha Maʼlumotlar"
            
            # Sarlavhalar
            headers = [
                'Sana', 'Vaqt', 'Tashkilot Nomi', 'Avtomobil Nomi', 
                'Avtomobil Raqami', 'Yoqilgʻi Miqdori (L)', 'Yoqilgʻi Turi',
                'Yoqilgʻi Narxi (soʻm)', 'Jami Narx (soʻm)', 'Foydalanuvchi', 'Rasm Linki'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            # Ma'lumotlarni qo'shish
            row_num = 2
            for record in compilated_data.select_related('tashkilot', 'avto', 'who_user'):
                # Yoqilg'i narxini hisoblash (Decimal va float muammosini hal qilish)
                try:
                    if record.hajm > 0:
                        # Ikkala qiymatni ham float ga o'tkazish
                        fuel_price_per_liter = float(record.all_price) / float(record.hajm)
                    else:
                        fuel_price_per_liter = 0
                except (TypeError, ValueError):
                    fuel_price_per_liter = 0
                
                process_photo_url = f"{request.scheme}://{request.get_host()}{record.photo.url}" if record.photo and record.photo.name != "none.jpg" else "Rasm mavjud emas"
                
                data_row = [
                    record.created_ad.strftime('%Y-%m-%d'),
                    record.created_ad.strftime('%H:%M'),
                    record.tashkilot.title,
                    record.avto.title,
                    record.avto.avto_number,
                    float(record.hajm),  # Float ga o'tkazish
                    record.yoqilgi_turi,
                    f"{fuel_price_per_liter:.0f}",
                    f"{float(record.all_price):.0f}",  # Float ga o'tkazish
                    record.who_user.username,
                    process_photo_url
                ]
                
                for col, value in enumerate(data_row, 1):
                    ws_main.cell(row=row_num, column=col, value=value)
                
                row_num += 1
            
            # Ustun enlarini sozlash
            column_widths = [12, 10, 20, 20, 15, 18, 15, 18, 15, 15, 30]
            for i, width in enumerate(column_widths, 1):
                ws_main.column_dimensions[get_column_letter(i)].width = width
            
            # 2. ASOSIY STATISTIKA varaqasi
            ws_stats = wb.create_sheet("Asosiy Statistika")
            
            # Aggregatlarni float ga o'tkazish
            total_fuel_result = compilated_data.aggregate(total=Sum('hajm'))
            total_price_result = compilated_data.aggregate(total=Sum('all_price'))
            
            total_fuel = float(total_fuel_result['total'] or 0)
            total_price = float(total_price_result['total'] or 0)
            
            # Sarlavhalar
            ws_stats.cell(row=1, column=1, value='Koʻrsatkich').fill = header_fill
            ws_stats.cell(row=1, column=1).font = header_font
            ws_stats.cell(row=1, column=2, value='Qiymat').fill = header_fill
            ws_stats.cell(row=1, column=2).font = header_font
            
            stats_data = [
                ('Umumiy Yoqilgʻi Miqdori', f"{total_fuel:.1f} L"),
                ('Jami Sarflangan Mablagʻ', f"{total_price:,.0f} soʻm".replace(',', ' ')),
                ('Jami Yozuvlar Soni', compilated_data.count()),
                ('Faol Avtomobillar', compilated_data.values('avto').distinct().count()),
                ('Faol Tashkilotlar', compilated_data.values('tashkilot').distinct().count()),
                ('Faol Foydalanuvchilar', compilated_data.values('who_user').distinct().count()),
                ('Kunlik Oʻrtacha Yoqilgʻi', f"{total_fuel / period_days:.1f} L"),
                ('Kunlik Oʻrtacha Mablagʻ', f"{total_price / period_days:,.0f} soʻm".replace(',', ' ')),
                ('Maksimal Yoqilgʻi Miqdori', f"{float(compilated_data.aggregate(max=Max('hajm'))['max'] or 0):.1f} L"),
                ('Minimal Yoqilgʻi Miqdori', f"{float(compilated_data.aggregate(min=Min('hajm'))['min'] or 0):.1f} L"),
                ('Davr', f"So'nggi {period_days} kun")
            ]
            
            for row, (indicator, value) in enumerate(stats_data, 2):
                ws_stats.cell(row=row, column=1, value=indicator)
                ws_stats.cell(row=row, column=2, value=value)
            
            ws_stats.column_dimensions['A'].width = 30
            ws_stats.column_dimensions['B'].width = 25
            
            # 3. TASHKILOTLAR STATISTIKASI
            ws_tashkilot = wb.create_sheet("Tashkilotlar Statistika")
            
            tashkilot_stats = compilated_data.values(
                'tashkilot__title'
            ).annotate(
                total_fuel=Sum('hajm'),
                total_price=Sum('all_price'),
                record_count=Count('id'),
                avg_fuel=Avg('hajm'),
                avto_count=Count('avto', distinct=True)
            ).order_by('-total_fuel')
            
            tashkilot_headers = [
                'Tashkilot Nomi', 'Umumiy Yoqilgʻi (L)', 'Jami Sarf (soʻm)', 
                'Yozuvlar Soni', 'Avtomobillar Soni', 'Oʻrtacha Yoqilgʻi (L)'
            ]
            
            for col, header in enumerate(tashkilot_headers, 1):
                cell = ws_tashkilot.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            row_num = 2
            for stat in tashkilot_stats:
                ws_tashkilot.cell(row=row_num, column=1, value=stat['tashkilot__title'])
                ws_tashkilot.cell(row=row_num, column=2, value=float(stat['total_fuel'] or 0))
                ws_tashkilot.cell(row=row_num, column=3, value=f"{float(stat['total_price'] or 0):,.0f}".replace(',', ' '))
                ws_tashkilot.cell(row=row_num, column=4, value=stat['record_count'])
                ws_tashkilot.cell(row=row_num, column=5, value=stat['avto_count'])
                ws_tashkilot.cell(row=row_num, column=6, value=float(stat['avg_fuel'] or 0))
                row_num += 1
            
            tashkilot_widths = [25, 20, 20, 15, 18, 20]
            for i, width in enumerate(tashkilot_widths, 1):
                ws_tashkilot.column_dimensions[get_column_letter(i)].width = width
            
            # 4. AVTOMOBILLAR STATISTIKASI
            ws_avto = wb.create_sheet("Avtomobillar Statistika")
            
            avto_stats = compilated_data.values(
                'avto__title', 'avto__avto_number', 'tashkilot__title'
            ).annotate(
                total_fuel=Sum('hajm'),
                total_price=Sum('all_price'),
                record_count=Count('id'),
                avg_fuel=Avg('hajm'),
                last_refuel=Max('created_ad')
            ).order_by('-total_fuel')
            
            avto_headers = [
                'Avtomobil Nomi', 'Avtomobil Raqami', 'Tashkilot',
                'Umumiy Yoqilgʻi (L)', 'Jami Sarf (soʻm)', 'Yozuvlar Soni',
                'Oʻrtacha Yoqilgʻi (L)', 'Oxirgi Marta'
            ]
            
            for col, header in enumerate(avto_headers, 1):
                cell = ws_avto.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            row_num = 2
            for stat in avto_stats:
                ws_avto.cell(row=row_num, column=1, value=stat['avto__title'])
                ws_avto.cell(row=row_num, column=2, value=stat['avto__avto_number'])
                ws_avto.cell(row=row_num, column=3, value=stat['tashkilot__title'])
                ws_avto.cell(row=row_num, column=4, value=float(stat['total_fuel'] or 0))
                ws_avto.cell(row=row_num, column=5, value=f"{float(stat['total_price'] or 0):,.0f}".replace(',', ' '))
                ws_avto.cell(row=row_num, column=6, value=stat['record_count'])
                ws_avto.cell(row=row_num, column=7, value=float(stat['avg_fuel'] or 0))
                ws_avto.cell(row=row_num, column=8, value=stat['last_refuel'].strftime('%Y-%m-%d %H:%M'))
                row_num += 1
            
            avto_widths = [20, 15, 20, 20, 20, 15, 20, 18]
            for i, width in enumerate(avto_widths, 1):
                ws_avto.column_dimensions[get_column_letter(i)].width = width

        # MUHIM: Faol varaqani belgilash (xatoni oldini olish uchun)
        wb.active = 0
        
        # Faylni saqlash
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"yoqilgi_statistikasi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel yaratish xatosi: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': str(e)
        })
@login_required
def export_today_excel(request):
    """Bugungi ma'lumotlarni Excel faylga eksport qilish"""
    try:
        today = timezone.now().date()
        
        # Filtrlarni olish
        tashkilot_id = request.GET.get('tashkilot_id', 'all')
        
        # Bugungi ma'lumotlarni olish
        compilated_data = Compilated.objects.filter(created_ad__date=today)
        
        if tashkilot_id != 'all':
            compilated_data = compilated_data.filter(tashkilot_id=tashkilot_id)
        
        # Ma'lumotlarni tayyorlash
        data = compilated_data.select_related('tashkilot', 'avto', 'who_user').order_by('-created_ad')
        
        # Excel fayl yaratish
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Bugungi barcha yozuvlar
            main_data = []
            for record in data:
                main_data.append({
                    'Vaqt': record.created_ad.strftime('%H:%M'),
                    'Tashkilot': record.tashkilot.title,
                    'Avtomobil': record.avto.title,
                    'Avtomobil Raqami': record.avto.avto_number,
                    'Yoqilgʻi Turi': record.yoqilgi_turi,
                    'Yoqilgʻi Miqdori (L)': record.hajm,
                    'Foydalanuvchi': record.who_user.username,
                    'Rasm Fayli': record.photo.name if record.photo else 'Mavjud emas'
                })
            
            df_main = pd.DataFrame(main_data)
            df_main.to_excel(writer, sheet_name="Bugungi Maʼlumotlar", index=False)
            
            # 2. Bugungi statistika
            total_fuel = data.aggregate(total=Sum('hajm'))['total'] or 0
            stats_data = {
                'Umumiy Yoqilgʻi (L)': [round(total_fuel, 1)],
                'Jami Yozuvlar': [data.count()],
                'Faol Avtomobillar': [data.values('avto').distinct().count()],
                'Faol Tashkilotlar': [data.values('tashkilot').distinct().count()],
                'Maksimal Yoqilgʻi (L)': [round(data.aggregate(max=Max('hajm'))['max'] or 0, 1)],
                'Minimal Yoqilgʻi (L)': [round(data.aggregate(min=Min('hajm'))['min'] or 0, 1)],
                'Oʻrtacha Yoqilgʻi (L)': [round(data.aggregate(avg=Avg('hajm'))['avg'] or 0, 1)]
            }
            
            df_stats = pd.DataFrame(stats_data)
            df_stats.to_excel(writer, sheet_name='Bugungi Statistika', index=False)
            
            # 3. Tashkilotlar bo'yicha bugungi statistika
            tashkilot_stats = data.values(
                'tashkilot__title'
            ).annotate(
                total_fuel=Sum('hajm'),
                record_count=Count('id')
            ).order_by('-total_fuel')
            
            tashkilot_data = []
            for stat in tashkilot_stats:
                tashkilot_data.append({
                    'Tashkilot': stat['tashkilot__title'],
                    'Umumiy Yoqilgʻi (L)': round(stat['total_fuel'], 1),
                    'Yozuvlar Soni': stat['record_count']
                })
            
            df_tashkilot = pd.DataFrame(tashkilot_data)
            df_tashkilot.to_excel(writer, sheet_name='Tashkilotlar', index=False)
            
            # 4. Soatlik statistika
            hourly_stats = data.annotate(
                hour=ExtractHour('created_ad')
            ).values('hour').annotate(
                hourly_total=Sum('hajm'),
                record_count=Count('id')
            ).order_by('hour')
            
            hourly_data = []
            for stat in hourly_stats:
                hourly_data.append({
                    'Soat': f"{stat['hour']}:00-{stat['hour']}:59",
                    'Yoqilgʻi Miqdori (L)': round(stat['hourly_total'], 1),
                    'Yozuvlar Soni': stat['record_count']
                })
            
            df_hourly = pd.DataFrame(hourly_data)
            df_hourly.to_excel(writer, sheet_name='Soatlik Statistika', index=False)
        
        # HTTP response yaratish
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Fayl nomi
        filename = f"bugungi_yoqilgi_{today.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Tashkilot, Avto, Yoqilgi_turi
from .forms import TashkilotForm, AvtoForm, YoqilgiTuriForm
@login_required
def admin_panel(request):
    context = {
        "users":User.objects.filter(is_superuser=False),
        'tashkilotlar': Tashkilot.objects.all(),
        'avto_list': Avto.objects.select_related('tashkilot').all(),
        "avtos":Avto.objects.all(),
        'yoqilgilar': Yoqilgi_turi.objects.all(),
    }
    return render(request, 'admin.html', context)

# Tashkilot CRUD
@login_required
def tashkilot_edit(request, pk):
    tashkilot = get_object_or_404(Tashkilot, pk=pk)
    if request.method == 'POST':
        form = TashkilotForm(request.POST, instance=tashkilot)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tashkilot muvaffaqiyatli yangilandi!')
            return redirect('admin_panel')
    else:
        form = TashkilotForm(instance=tashkilot)
    return render(request, 'tashkilot_edit.html', {'form': form})
@login_required
def tashkilot_delete(request, pk):
    tashkilot = get_object_or_404(Tashkilot, pk=pk)
    if request.method == 'POST':
        tashkilot.delete()
        messages.success(request, 'Tashkilot muvaffaqiyatli o\'chirildi!')
        return redirect('admin_panel')
    return render(request, 'tashkilot_delete.html', {'tashkilot': tashkilot})

# Avto CRUD (xuddi shu logika)
@login_required
def avto_edit(request, pk):
    avto = get_object_or_404(Avto, pk=pk)
    tashkilotlar = Tashkilot.objects.all()
    if request.method == 'POST':
        form = AvtoForm(request.POST, instance=avto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Avtomobil muvaffaqiyatli yangilandi!')
            return redirect('admin_panel')
    else:
        form = AvtoForm(instance=avto)
    return render(request, 'avto_edit.html', {'form': form,"tashkilotlar":tashkilotlar})
@login_required
def avto_delete(request, pk):
    avto = get_object_or_404(Avto, pk=pk)
    if request.method == 'POST':
        avto.delete()
        messages.success(request, 'Avtomobil muvaffaqiyatli o\'chirildi!')
        return redirect('admin_panel')
    return render(request, 'avto_delete.html', {'avto': avto})

# Yoqilgi turi CRUD (xuddi shu logika)
@login_required
def yoqilgi_turi_edit(request, pk):
    yoqilgi = get_object_or_404(Yoqilgi_turi, pk=pk)
    if request.method == 'POST':
        form = YoqilgiTuriForm(request.POST, instance=yoqilgi)
        if form.is_valid():
            form.save()
            messages.success(request, 'Yoqilg\'i turi muvaffaqiyatli yangilandi!')
            return redirect('admin_panel')
    else:
        form = YoqilgiTuriForm(instance=yoqilgi)
    return render(request, 'yoqilgi_turi_edit.html', {'form': form})
@login_required
def yoqilgi_turi_delete(request, pk):
    yoqilgi = get_object_or_404(Yoqilgi_turi, pk=pk)
    if request.method == 'POST':
        yoqilgi.delete()
        messages.success(request, 'Yoqilg\'i turi muvaffaqiyatli o\'chirildi!')
        return redirect('admin_panel')
    return render(request, 'yoqilgi_turi_delete.html', {'yoqilgi_turi': yoqilgi})


@login_required
def home_worker(request):
    """Oddiy ishchi uchun bosh sahifa"""
    try:
        # Bugungi sana
        today = timezone.now().date()
        
        # Bugungi ma'lumotlar
        today_records = Compilated.objects.filter(created_ad__date=today)
        total_fuel = today_records.aggregate(total=Sum('hajm'))['total'] or 0
        total_count = today_records.count()
        active_cars = today_records.values('avto').distinct().count()
        avg_fuel = round(total_fuel / active_cars, 1) if active_cars > 0 else 0
        
        # So'nggi faoliyat
        recent_activities = today_records.select_related('tashkilot', 'avto').order_by('-created_ad')[:5]
        
        # Tashkilotlar ro'yxati (avtomobil qo'shish uchun)
        tashkilotlar = Tashkilot.objects.all()
        
        # So'nggi faoliyatni formatlash
        recent_activities_list = []
        for activity in recent_activities:
            recent_activities_list.append({
                'avto_title': activity.avto.title if activity.avto else 'Noma\'lum',
                'avto_number': activity.avto.avto_number if activity.avto else '',
                'tashkilot_title': activity.tashkilot.title if activity.tashkilot else 'Noma\'lum',
                'hajm': activity.hajm,
                'time': activity.created_ad.strftime('%H:%M')
            })
        
        context = {
            'total_fuel': round(total_fuel, 1),
            'total_count': total_count,
            'active_cars': active_cars,
            'avg_fuel': avg_fuel,
            'recent_activities': recent_activities_list,
            'tashkilotlar': tashkilotlar,
        }
        
        return render(request, 'home_worker.html', context)
        
    except Exception as e:
        # Xatolik yuz bersa, bo'sh ma'lumotlar bilan sahifani ko'rsatish
        context = {
            'total_fuel': 0,
            'total_count': 0,
            'active_cars': 0,
            'avg_fuel': 0,
            'recent_activities': [],
            'tashkilotlar': Tashkilot.objects.all(),
        }
        return render(request, 'home_worker.html', context)

@login_required
def add_tashkilot_worker(request):
    """Oddiy ishchi uchun tashkilot qo'shish"""
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            group_id = request.POST.get('group_id', 4885110792)
            
            if not title:
                messages.error(request, 'Tashkilot nomini kiriting!')
                return redirect('home')
            
            # Tashkilotni yaratish
            tashkilot = Tashkilot.objects.create(
                title=title,
                group_id=group_id
            )
            
            messages.success(request, f'"{title}" tashkiloti muvaffaqiyatli qo\'shildi!')
            
        except Exception as e:
            messages.error(request, f'Tashkilot qo\'shishda xatolik: {str(e)}')
    
    return redirect('home')

@login_required
def add_avto_worker(request):
    """Oddiy ishchi uchun avtomobil qo'shish"""
    if request.method == 'POST':
        try:
            tashkilot_id = request.POST.get('tashkilot_id')
            title = request.POST.get('title')
            avto_number = request.POST.get('avto_number')
            
            if not all([tashkilot_id, title, avto_number]):
                messages.error(request, 'Barcha maydonlarni to\'ldiring!')
                return redirect('home')
            
            # Tashkilotni tekshirish
            try:
                tashkilot = Tashkilot.objects.get(id=tashkilot_id)
            except Tashkilot.DoesNotExist:
                messages.error(request, 'Tashkilot topilmadi!')
                return redirect('home')
            
            # Avtomobilni yaratish
            avto = Avto.objects.create(
                tashkilot=tashkilot,
                title=title,
                avto_number=avto_number
            )
            
            messages.success(request, f'"{title}" avtomobili muvaffaqiyatli qo\'shildi!')
            
        except Exception as e:
            messages.error(request, f'Avtomobil qo\'shishda xatolik: {str(e)}')
    
    return redirect('home')